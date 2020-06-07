import pandas as pd
from os import listdir
from os.path import isfile, isdir, join
from openpyxl import load_workbook
import xlsxwriter

# Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook('D:\\python_code\\export_dataframe.xlsx')
worksheet = workbook.add_worksheet()
workbook.close()

#load exist file and
book = load_workbook('D:\\python_code\\export_dataframe.xlsx')
writer = pd.ExcelWriter('D:\\python_code\\export_dataframe.xlsx', engine='openpyxl')
#writer.book = book

#record the file in the directory
mypath = r"D:\python_code\file"
files = listdir(mypath)
list_files=[]
for f in files:
  fullpath = join(mypath, f)
  if isfile(fullpath):
    print("檔案：", f)
    list_files.append(f)
  elif isdir(fullpath):
    print("目錄：", f)
print(list_files)
#read the file and extract the data we need from pandas dataframe
for i in range(len(list_files)):
  f = open(mypath+"\\"+list_files[i],'r')
  table = pd.read_table(f,sep=";")
  infor = table["askldfjaslkf"]
  infor.to_excel(writer, startcol= i,index = False, header=True)

writer.save()
f.close()